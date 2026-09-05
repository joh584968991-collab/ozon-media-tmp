#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""为单个目标 SKU 生成可审计的 Ozon 轮播视频。

成功与否只以 ``RESULT_JSON`` 和退出码为准：
0 = CDN_READY（或 --no-push 的 GENERATED_COMPLIANT）
2 = PUSHED_CDN_PENDING，已推送但 CDN 还不能确认可用
1 = FAILED
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import re
import subprocess
import sys
import tempfile
from dataclasses import dataclass, field
from pathlib import Path
from urllib.parse import quote, urlparse

XLSX = Path('/Users/capper/.openclaw/workspace-ozon-lister/workbench/ozon8_final3_titles_fixed.xlsx')
COMBO_COMPONENTS_TABLE = Path('/Users/capper/.openclaw/workspace-ozon-lister/workbench/cost/组合采购价表.txt')
OUT_BASE = Path('/Users/capper/.openclaw/workspace-ozon-media')
REPO = OUT_BASE
IMAGE_SOURCES_CONFIG_PATH = OUT_BASE / 'image_sources.json'

# 路径安全和文件命名只允许真实货号常用的 ASCII 字符；支持如 QK1923-L 的后缀。
SKU_PATTERN = re.compile(r'^[A-Za-z0-9][A-Za-z0-9_-]{0,63}$')
C_SKU = 2
C_MAIN = 12
C_EXTRA0 = 20
MAX_IMAGES = 9
MIN_IMAGES = 4  # 每图 2 秒，Ozon 最短视频为 8 秒。


class PipelineError(RuntimeError):
    """表示可向 Agent 明确报告的流水线失败。"""


class SourceSkuNotFoundError(PipelineError):
    """仅表示 xlsx 中没有请求 SKU；其他源数据错误不得触发组合回退。"""


@dataclass(frozen=True)
class ImageSource:
    """图片来源的可审计证明；目标 SKU 与图片来源 SKU 可以不同。"""

    target_sku: str
    source_sku: str
    image_urls: list[str]
    mode: str
    component_quantity: int | None = None
    mapping_file: str | None = None

    def evidence(self) -> dict[str, object]:
        result: dict[str, object] = {
            'mode': self.mode,
            'target_sku': self.target_sku,
            'source_sku': self.source_sku,
            'source_url_count': len(self.image_urls),
        }
        if self.component_quantity is not None:
            result['component_quantity'] = self.component_quantity
        if self.mapping_file is not None:
            result['mapping_file'] = self.mapping_file
        return result


def emit_result(status: str, **details: object) -> None:
    """输出供 Agent 逐字段读取的最终状态，不把候选值说成已验证事实。"""
    print('RESULT_JSON=' + json.dumps({'status': status, **details}, ensure_ascii=False, sort_keys=True))


def normalize_sku(value: str) -> str:
    sku = value.strip()
    if not SKU_PATTERN.fullmatch(sku):
        raise PipelineError('SKU 格式非法：仅允许字母、数字、连字符和下划线，且最长 64 个字符')
    return sku


def is_http_url(value: str) -> bool:
    parsed = urlparse(value)
    return parsed.scheme in {'http', 'https'} and bool(parsed.netloc)


def get_images_from_xlsx(sku_id: str) -> list[str]:
    """只从源 xlsx 的唯一同名 SKU 行读取主图和附加图。"""
    if not XLSX.is_file():
        raise PipelineError(f'源图片表不存在：{XLSX}')

    try:
        import openpyxl
    except ImportError as exc:
        raise PipelineError('缺少 openpyxl，无法读取源图片表') from exc

    try:
        workbook = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    except Exception as exc:
        raise PipelineError(f'无法打开源图片表：{exc}') from exc

    try:
        if 'Sheet1' not in workbook.sheetnames:
            raise PipelineError('源图片表缺少 Sheet1 工作表')
        sheet = workbook['Sheet1']
        matches: list[list[str]] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) <= C_SKU or row[C_SKU] in (None, ''):
                continue
            if str(row[C_SKU]).strip() != sku_id:
                continue
            raw_values = [row[C_MAIN] if len(row) > C_MAIN else None]
            raw_values.extend(
                row[column] if len(row) > column else None
                for column in range(C_EXTRA0, C_EXTRA0 + 14)
            )
            matches.append([
                str(value).strip()
                for value in raw_values
                if value not in (None, '') and is_http_url(str(value).strip())
            ])
    finally:
        workbook.close()

    if not matches:
        raise SourceSkuNotFoundError(f'源图片表中未找到 SKU {sku_id}')
    if len(matches) != 1:
        raise PipelineError(f'源图片表中 SKU {sku_id} 出现 {len(matches)} 次，拒绝选择其中一行')

    unique_urls: list[str] = []
    seen: set[str] = set()
    for url in matches[0]:
        if url not in seen:
            unique_urls.append(url)
            seen.add(url)
    return unique_urls


def load_combo_components() -> dict[str, list[tuple[str, int]]]:
    """从权威组合采购价表读取 ``组合 SKU -> 单品 SKU × 数量`` 关系。"""
    if not COMBO_COMPONENTS_TABLE.is_file():
        raise PipelineError(f'组合 SKU 映射表不存在：{COMBO_COMPONENTS_TABLE}')

    try:
        handle = COMBO_COMPONENTS_TABLE.open(encoding='utf-8-sig', newline='')
    except OSError as exc:
        raise PipelineError(f'无法读取组合 SKU 映射表：{exc}') from exc

    with handle:
        reader = csv.DictReader(handle, delimiter='\t')
        required_columns = {'组合SKU', '单品SKU', '单品数量'}
        if not reader.fieldnames or not required_columns.issubset(reader.fieldnames):
            raise PipelineError('组合 SKU 映射表缺少 组合SKU / 单品SKU / 单品数量 列')

        components: dict[str, list[tuple[str, int]]] = {}
        current_combo: str | None = None
        for line_number, row in enumerate(reader, start=2):
            combo_value = (row.get('组合SKU') or '').strip()
            component_value = (row.get('单品SKU') or '').strip()
            quantity_value = (row.get('单品数量') or '').strip()
            if combo_value:
                current_combo = normalize_sku(combo_value)
                components.setdefault(current_combo, [])
            if not component_value and not quantity_value:
                continue
            if current_combo is None:
                raise PipelineError(f'组合 SKU 映射表第 {line_number} 行没有组合 SKU 上下文')
            if not component_value or not quantity_value:
                raise PipelineError(f'组合 SKU 映射表第 {line_number} 行的单品 SKU 或数量为空')
            try:
                quantity = int(quantity_value)
            except ValueError as exc:
                raise PipelineError(f'组合 SKU 映射表第 {line_number} 行数量不是正整数：{quantity_value}') from exc
            if quantity <= 0:
                raise PipelineError(f'组合 SKU 映射表第 {line_number} 行数量必须大于 0')
            components[current_combo].append((normalize_sku(component_value), quantity))
    # 组合表偶有只有组合 SKU、没有任何单品行的占位记录；它们不是可用映射，必须保持阻断。
    return {combo_sku: entries for combo_sku, entries in components.items() if entries}


@dataclass(frozen=True)
class ImageSourceConfig:
    """可插拔的图片来源后端声明。

    - ``sku_pattern`` 为 None 表示 catch-all 默认后端（如 xlsx）。
    - ``priority`` 升序遍历，越小越优先；多个后端可声明不同 SKU 段。
    """
    name: str
    kind: str
    priority: int = 100
    sku_pattern: str | None = None
    options: dict = field(default_factory=dict)

    def matches(self, sku: str) -> bool:
        if self.sku_pattern is None:
            return True
        return bool(re.fullmatch(self.sku_pattern, sku))


class ImageSourceBackend:
    """图片来源后端协议：把任意存储（xlsx / 远端仓库 / 本地目录）映射为可下载 URL 列表。"""
    name: str = ''

    def resolve(self, target_sku: str) -> ImageSource:
        raise NotImplementedError

    def describe(self) -> dict[str, object]:
        return {'name': self.name}


class XlsxImageBackend(ImageSourceBackend):
    """从源 xlsx + 组合采购价表读取图片，保留原 DIRECT / DERIVED_SINGLE_COMPONENT 语义。"""
    name = 'xlsx'

    def __init__(self, options: dict | None = None):
        self.options = options or {}

    def resolve(self, target_sku: str) -> ImageSource:
        try:
            direct_urls = get_images_from_xlsx(target_sku)
        except SourceSkuNotFoundError:
            pass
        else:
            return ImageSource(
                target_sku=target_sku,
                source_sku=target_sku,
                image_urls=direct_urls,
                mode='DIRECT',
            )

        components = load_combo_components().get(target_sku)
        if not components:
            raise SourceSkuNotFoundError(
                f'xlsx 中未找到 SKU {target_sku}，组合采购价表也没有该 SKU'
            )
        if len(components) != 1:
            raise PipelineError(
                f'组合 SKU {target_sku} 含 {len(components)} 个单品组成，禁止用任一单品图片代替组合图片'
            )

        source_sku, component_quantity = components[0]
        try:
            source_urls = get_images_from_xlsx(source_sku)
        except SourceSkuNotFoundError as exc:
            raise PipelineError(
                f'组合 SKU {target_sku} 可回退到单品 {source_sku}，但该单品在 xlsx 中不存在'
            ) from exc
        return ImageSource(
            target_sku=target_sku,
            source_sku=source_sku,
            image_urls=source_urls,
            mode='DERIVED_SINGLE_COMPONENT',
            component_quantity=component_quantity,
            mapping_file=str(COMBO_COMPONENTS_TABLE),
        )

    def describe(self) -> dict[str, object]:
        return {'name': self.name, 'xlsx': str(XLSX), 'combo': str(COMBO_COMPONENTS_TABLE)}


class RemoteRepoImageBackend(ImageSourceBackend):
    """从远端 GitHub 公开仓库 raw 链接读取图片。

    通过 ``download_image`` 逐张 ffprobe 校验，确认真实可解码 PNG 才纳入 URL 列表；
    假定文件名按 ``{nn:02d}.png`` 从 min_index 连续递增，缺一张即停止以避免混入违规素材。
    """
    name = 'remote_repo'

    def __init__(self, owner: str, repo: str, sha: str,
                 path_template: str = 'images/{sku}/{nn:02d}.png',
                 min_index: int = 1, max_index: int = 9,
                 config_path: str = ''):
        self.base = f'https://raw.githubusercontent.com/{owner}/{repo}/{sha}'
        self.path_template = path_template
        self.min_index = min_index
        self.max_index = max_index
        self.config_path = config_path

    def _url(self, sku: str, nn: int) -> str:
        return f"{self.base}/{self.path_template.format(sku=sku, nn=nn)}"

    def resolve(self, target_sku: str) -> ImageSource:
        urls: list[str] = []
        with tempfile.TemporaryDirectory(prefix=f'.probe_{target_sku}_', dir=OUT_BASE) as tmp_name:
            tmp_dir = Path(tmp_name)
            for nn in range(self.min_index, self.max_index + 1):
                url = self._url(target_sku, nn)
                probe_file = tmp_dir / f'{nn:02d}.png'
                if download_image(url, probe_file):
                    urls.append(url)
                else:
                    break  # 连续 01..N；缺一张即停止

        if len(urls) < MIN_IMAGES:
            raise SourceSkuNotFoundError(
                f'远端仓库 {self.base} 找不到 SKU {target_sku} 的 ≥{MIN_IMAGES} 张图（实际 {len(urls)} 张）'
            )

        return ImageSource(
            target_sku=target_sku,
            source_sku=target_sku,
            image_urls=urls,
            mode='REMOTE_REPO',
            mapping_file=self.config_path,
        )

    def describe(self) -> dict[str, object]:
        return {
            'name': self.name,
            'base': self.base,
            'path_template': self.path_template,
            'min_index': self.min_index,
            'max_index': self.max_index,
        }


class LocalDirImageBackend(ImageSourceBackend):
    """从本地目录读取图片，按文件名递增序列拼接。"""
    name = 'local_dir'

    def __init__(self, base_dir: str,
                 sku_subdir: str = '{sku}',
                 filename_pattern: str = '{nn:02d}.png',
                 min_index: int = 1, max_index: int = 9,
                 config_path: str = ''):
        self.base_dir = Path(base_dir)
        self.sku_subdir = sku_subdir
        self.filename_pattern = filename_pattern
        self.min_index = min_index
        self.max_index = max_index
        self.config_path = config_path

    def resolve(self, target_sku: str) -> ImageSource:
        sub = self.base_dir / self.sku_subdir.format(sku=target_sku)
        if not sub.is_dir():
            raise SourceSkuNotFoundError(f'本地目录 {sub} 不存在')

        urls: list[str] = []
        for nn in range(self.min_index, self.max_index + 1):
            file_path = sub / self.filename_pattern.format(sku=target_sku, nn=f'{nn:02d}')
            if file_path.is_file() and image_is_decodable(file_path):
                urls.append(file_path.absolute().as_uri())
            else:
                break

        if len(urls) < MIN_IMAGES:
            raise SourceSkuNotFoundError(
                f'本地目录 {sub} 找不到 SKU {target_sku} 的 ≥{MIN_IMAGES} 张图（实际 {len(urls)} 张）'
            )

        return ImageSource(
            target_sku=target_sku,
            source_sku=target_sku,
            image_urls=urls,
            mode='LOCAL_DIR',
            mapping_file=self.config_path,
        )

    def describe(self) -> dict[str, object]:
        return {'name': self.name, 'base_dir': str(self.base_dir)}


def make_backend(cfg: ImageSourceConfig) -> ImageSourceBackend:
    """根据配置构造后端实例；未知 kind 抛 PipelineError。"""
    if cfg.kind == 'xlsx':
        return XlsxImageBackend(cfg.options)
    if cfg.kind == 'remote_repo':
        opts = cfg.options
        for required_key in ('owner', 'repo', 'sha'):
            if required_key not in opts:
                raise PipelineError(f'remote_repo 配置缺少字段 {required_key!r}')
        return RemoteRepoImageBackend(
            owner=opts['owner'],
            repo=opts['repo'],
            sha=opts['sha'],
            path_template=opts.get('path_template', 'images/{sku}/{nn:02d}.png'),
            min_index=int(opts.get('min_index', 1)),
            max_index=int(opts.get('max_index', 9)),
            config_path=str(IMAGE_SOURCES_CONFIG_PATH),
        )
    if cfg.kind == 'local_dir':
        opts = cfg.options
        if 'base_dir' not in opts:
            raise PipelineError('local_dir 配置缺少字段 base_dir')
        return LocalDirImageBackend(
            base_dir=opts['base_dir'],
            sku_subdir=opts.get('sku_subdir', '{sku}'),
            filename_pattern=opts.get('filename_pattern', '{nn:02d}.png'),
            min_index=int(opts.get('min_index', 1)),
            max_index=int(opts.get('max_index', 9)),
            config_path=str(IMAGE_SOURCES_CONFIG_PATH),
        )
    raise PipelineError(f'未知图片来源后端类型: {cfg.kind!r}')


def load_image_source_config(path: Path | None = None) -> list[ImageSourceConfig]:
    """读 image_sources.json；缺省时只启用 xlsx 后端（保持旧行为完全兼容）。"""
    cfg_path = path or IMAGE_SOURCES_CONFIG_PATH
    if not cfg_path.is_file():
        return [ImageSourceConfig(name='xlsx_primary', kind='xlsx', priority=100, sku_pattern=None, options={})]
    try:
        data = json.loads(cfg_path.read_text(encoding='utf-8'))
    except json.JSONDecodeError as exc:
        raise PipelineError(f'image_sources.json 解析失败：{exc}') from exc
    sources_raw = data.get('sources')
    if not isinstance(sources_raw, list) or not sources_raw:
        raise PipelineError('image_sources.json 必须包含非空 sources 数组')
    configs: list[ImageSourceConfig] = []
    for entry in sources_raw:
        if not isinstance(entry, dict):
            raise PipelineError('image_sources.json 的 sources 每项必须为对象')
        if 'kind' not in entry:
            raise PipelineError('image_sources.json 的 sources 每项必须包含 kind')
        cfg = ImageSourceConfig(
            name=str(entry.get('name') or entry['kind']),
            kind=str(entry['kind']),
            priority=int(entry.get('priority', 100)),
            sku_pattern=entry.get('sku_pattern'),
            options=entry.get('options') or {},
        )
        configs.append(cfg)
    return configs


def resolve_image_source(target_sku: str,
                         config: list[ImageSourceConfig] | None = None,
                         ) -> tuple[ImageSource, str]:
    """按 priority 升序遍历已声明的图片来源后端。

    对 SourceSkuNotFoundError 宽容（继续试下一个后端），其他 PipelineError 立即抛。
    返回 ``(ImageSource, 后端名)`` 以便 RESULT_JSON 标注 backend 与 backend_config。
    """
    configs = config if config is not None else load_image_source_config()
    sorted_configs = sorted(configs, key=lambda c: c.priority)
    last_err: str | None = None
    matched_any = False
    for cfg in sorted_configs:
        if not cfg.matches(target_sku):
            continue
        matched_any = True
        backend = make_backend(cfg)
        try:
            image_source = backend.resolve(target_sku)
            return image_source, backend.name
        except SourceSkuNotFoundError as exc:
            last_err = f'{backend.name}: {exc}'
            continue
    if not matched_any:
        raise PipelineError(f'没有任何图片来源后端匹配 SKU {target_sku}')
    raise PipelineError(f'所有匹配的后端均找不到 SKU {target_sku}：{last_err or "unknown"}')


def image_is_decodable(path: Path) -> bool:
    """以 ffprobe 检验下载内容确实是可解码图像，而不是 HTML 错误页。"""
    result = subprocess.run(
        [
            'ffprobe', '-v', 'error', '-select_streams', 'v:0',
            '-show_entries', 'stream=codec_type,width,height', '-of', 'json', str(path),
        ],
        capture_output=True,
        text=True,
    )
    if result.returncode != 0:
        return False
    try:
        streams = json.loads(result.stdout).get('streams', [])
    except json.JSONDecodeError:
        return False
    if not streams:
        return False
    stream = streams[0]
    return (
        stream.get('codec_type') == 'video'
        and int(stream.get('width') or 0) > 0
        and int(stream.get('height') or 0) > 0
    )


def download_image(url: str, destination: Path) -> bool:
    """下载并校验单张源图；失败图片不会被伪装成可用输入。"""
    result = subprocess.run(
        ['curl', '-fsSL', '--connect-timeout', '10', '--max-time', '30', '-o', str(destination), url],
        capture_output=True,
        text=True,
    )
    if result.returncode != 0 or not destination.is_file() or destination.stat().st_size <= 1000:
        destination.unlink(missing_ok=True)
        return False
    if not image_is_decodable(destination):
        destination.unlink(missing_ok=True)
        return False
    return True


def build_filter(image_count: int) -> str:
    """构造明确限定为 BT.709 limited-range / yuv420p 的滤镜图。"""
    parts = []
    for index in range(image_count):
        parts.append(
            f'[{index}:v]scale=800:800:force_original_aspect_ratio=decrease:out_range=tv,'
            f'pad=800:800:(ow-iw)/2:(oh-ih)/2:white,setsar=1,format=yuv420p,'
            f'setparams=range=tv:color_primaries=bt709:color_trc=bt709:colorspace=bt709[v{index}]'
        )
    concat_inputs = ''.join(f'[v{index}]' for index in range(image_count))
    parts.append(
        f'{concat_inputs}concat=n={image_count}:v=1:a=0,'
        'setparams=range=tv:color_primaries=bt709:color_trc=bt709:colorspace=bt709[outv]'
    )
    return ';'.join(parts)


def make_video(
    sku_id: str,
    image_urls: list[str],
    out_dir: Path,
    replace_existing: bool,
) -> tuple[Path, dict[str, object]]:
    """生成候选视频；只有本地合规检查通过才原子写入正式文件名。"""
    from check_video_compliance import check

    out_dir.mkdir(parents=True, exist_ok=True)
    output = out_dir / f'carousel_{sku_id}.mp4'
    if output.exists() and not replace_existing:
        raise PipelineError(
            f'目标文件已存在：{output.name}；禁止静默覆盖。仅在用户明确确认后使用 --replace-existing'
        )

    with tempfile.TemporaryDirectory(prefix=f'.tmp_{sku_id}_', dir=out_dir) as tmp_name:
        tmp_dir = Path(tmp_name)
        image_files: list[Path] = []
        for index, url in enumerate(image_urls[:MAX_IMAGES], start=1):
            destination = tmp_dir / f'img{index}'
            if download_image(url, destination):
                image_files.append(destination)

        if len(image_files) < MIN_IMAGES:
            raise PipelineError(
                f'可解码图片不足 {MIN_IMAGES} 张（实际 {len(image_files)} 张），无法生成至少 8 秒的合规视频'
            )

        inputs: list[str] = []
        for image_file in image_files:
            inputs.extend(['-loop', '1', '-t', '2', '-i', str(image_file)])
        candidate = tmp_dir / output.name
        command = [
            'ffmpeg', '-y', *inputs,
            '-filter_complex', build_filter(len(image_files)),
            '-map', '[outv]', '-c:v', 'libx264', '-profile:v', 'main', '-pix_fmt', 'yuv420p',
            '-color_range', 'tv', '-movflags', '+faststart', str(candidate),
        ]
        result = subprocess.run(command, capture_output=True, text=True)
        if result.returncode != 0 or not candidate.is_file() or candidate.stat().st_size <= 10000:
            detail = (result.stderr or '没有 ffmpeg 错误输出').strip()[-500:]
            raise PipelineError(f'ffmpeg 生成失败：{detail}')

        issues, metadata = check(str(candidate))
        if issues:
            raise PipelineError('候选视频未通过合规检查：' + '；'.join(issues))
        metadata['image_count'] = len(image_files)

        # 候选文件和目标文件位于同一文件系统；失败时不会覆盖原资产。
        os.replace(candidate, output)
    return output, metadata


def run_git(repo: Path, args: list[str], env: dict[str, str] | None = None) -> str:
    result = subprocess.run(['git', '-C', str(repo), *args], capture_output=True, text=True, env=env)
    if result.returncode != 0:
        detail = (result.stderr or result.stdout or '没有 Git 错误输出').strip()[-700:]
        raise PipelineError(f'Git {args[0]} 失败：{detail}')
    return result.stdout.strip()


def push_to_github(video_path: Path, repo: Path = REPO) -> tuple[str, str]:
    """只提交当前视频，不碰用户已有暂存区；推送后核对远端分支提交。"""
    repo = repo.resolve()
    if video_path.resolve().parent != repo:
        raise PipelineError('仅允许推送仓库根目录中的视频文件')
    filename = video_path.name
    if not video_path.is_file() or not filename.startswith('carousel_') or not filename.endswith('.mp4'):
        raise PipelineError('待推送文件不存在或不符合 carousel_<SKU>.mp4 命名')

    descriptor, index_name = tempfile.mkstemp(prefix='.git-index-ozon-media-', dir=repo)
    os.close(descriptor)
    index_path = Path(index_name)
    index_path.unlink(missing_ok=True)
    env = os.environ.copy()
    env['GIT_INDEX_FILE'] = str(index_path)
    try:
        run_git(repo, ['read-tree', 'HEAD'], env)
        run_git(repo, ['add', '--', filename], env)
        staged = run_git(repo, ['diff', '--cached', '--name-only'], env).splitlines()
        if staged != [filename]:
            raise PipelineError(f'隔离提交包含预期外文件：{staged}')
        run_git(repo, ['commit', '-m', f'add {filename}'], env)
        commit = run_git(repo, ['rev-parse', 'HEAD'])
    finally:
        index_path.unlink(missing_ok=True)
        index_path.with_name(index_path.name + '.lock').unlink(missing_ok=True)

    run_git(repo, ['push', 'origin', 'HEAD:main'])
    remote_ref = run_git(repo, ['ls-remote', 'origin', 'refs/heads/main'])
    if not remote_ref.startswith(commit + '\t'):
        raise PipelineError('远端 main 未指向本次提交，不能宣布视频已推送')

    url = 'https://cdn.jsdelivr.net/gh/joh584968991-collab/ozon-media-tmp@main/' + quote(filename)
    return commit, url


def verify_cdn(url: str) -> tuple[int | None, str | None]:
    """仅在响应明确为 HTTP 200 且 video/mp4 时认定 CDN 已可用。"""
    result = subprocess.run(
        [
            'curl', '-sSIL', '--connect-timeout', '10', '--max-time', '20', '-o', '/dev/null',
            '-w', '%{http_code}\n%{content_type}', url,
        ],
        capture_output=True,
        text=True,
    )
    if result.returncode != 0:
        return None, None
    lines = result.stdout.splitlines()
    try:
        status_code = int(lines[0])
    except (IndexError, ValueError):
        status_code = None
    content_type = lines[1].strip() if len(lines) > 1 else ''
    return status_code, content_type or None


def main() -> int:
    parser = argparse.ArgumentParser(description='生成并受控推送 Ozon 轮播视频')
    parser.add_argument('sku', help='与源 xlsx 完全一致的 SKU')
    parser.add_argument('--no-push', action='store_true', help='仅生成并合规校验，用于本地验证')
    parser.add_argument('--out-dir', type=Path, default=OUT_BASE, help='输出目录；推送时必须为仓库根目录')
    parser.add_argument('--replace-existing', action='store_true', help='明确允许覆盖同名本地视频')
    args = parser.parse_args()

    stage = 'input'
    sku = args.sku.strip()
    try:
        sku = normalize_sku(args.sku)
        out_dir = args.out_dir.expanduser().resolve()
        if not args.no_push and out_dir != OUT_BASE.resolve():
            raise PipelineError('推送模式只能使用仓库根目录；自定义输出目录必须搭配 --no-push')

        image_source, image_backend = resolve_image_source(sku)
        image_evidence = dict(image_source.evidence())
        image_evidence['backend'] = image_backend
        stage = 'generation'
        video_path, metadata = make_video(sku, image_source.image_urls, out_dir, args.replace_existing)
        if args.no_push:
            emit_result(
                'GENERATED_COMPLIANT', sku=sku, video_path=str(video_path),
                image_source=image_evidence, compliance=metadata,
            )
            return 0

        stage = 'push'
        commit, candidate_url = push_to_github(video_path)
        stage = 'cdn'
        status_code, content_type = verify_cdn(candidate_url)
        content_type_base = (content_type or '').split(';', 1)[0].strip().lower()
        if status_code == 200 and content_type_base == 'video/mp4':
            emit_result(
                'CDN_READY', sku=sku, commit=commit, available_url=candidate_url,
                http_status=status_code, content_type=content_type, image_source=image_evidence,
            )
            return 0

        emit_result(
            'PUSHED_CDN_PENDING', sku=sku, commit=commit, candidate_url=candidate_url,
            http_status=status_code, content_type=content_type, image_source=image_evidence,
        )
        return 2
    except PipelineError as exc:
        emit_result('FAILED', sku=sku, stage=stage, error=str(exc))
        return 1
    except Exception as exc:  # 防止模型在脚本异常时从半截输出臆测成功。
        emit_result('FAILED', sku=sku, stage=stage, error=f'未处理异常：{type(exc).__name__}: {exc}')
        return 1


if __name__ == '__main__':
    sys.exit(main())
